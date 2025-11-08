# common.py
from openpyxl import load_workbook

FILE = "snack.xlsx"

def get_money(user):
    wb = load_workbook(FILE)
        ws = wb[user]
            val = ws["B2"].value
                wb.close()
                    return val

                    def set_money(user, value):
                        wb = load_workbook(FILE)
                            ws = wb[user]
                                ws["B2"] = value
                                    wb.save(FILE)
                                        wb.close()